import json
import os
from datasets import load_from_disk
import soundfile as sf
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_all_split(root_hf):
    directories = []
    for dirpath, dirs, files in os.walk(root_hf):
        if len(dirs) == 0:
            directories.append(dirpath)
    return directories


def generate(split):
    print(f"start {split}", flush=True)
    ds=load_from_disk(split)
    samples=[]
    for i in [1,11,21,31,41,51,61,71,81,91]:
        item=ds[i]
        save_path = split.replace("/datasets/", "/datasets/samples/")
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        sf.write("{}_{}_context.wav".format(save_path, i), item["context"]["audio"]["array"], 16000)
        if "answer" in item.keys() and item["answer"]["audio"]:
            sf.write("{}_{}_answer.wav".format(save_path, i), item["answer"]["audio"]["array"], 16000)
        with open("{}_{}.json".format(save_path, i), "w", encoding="utf8") as f_out:
            del item["context"]
            if "answer" in item.keys():
                item["answer"]["audio"] = None
            f_out.write(json.dumps(item, ensure_ascii=False, indent=2))
        with open("{}_{}.txt".format(save_path, i), "w", encoding="utf8") as f_txt:
            f_txt.write(item["answer"]["text"])
        samples.append(item["answer"]["text"])
    return samples


def json2excel(json_file):

    # Your JSON
    data = json.load(open(json_file))

    rows = []
    for dataset, sentences in data.items():
        for sentence in sentences:
            rows.append({"Dataset": dataset, "Text": sentence})

    df = pd.DataFrame(rows)

    # Step 1: Write DataFrame to Excel
    file_name = json_file.replace(".json", ".xlsx")
    df.to_excel(file_name, index=False)

    # Step 2: Open with openpyxl and merge dataset cells
    wb = load_workbook(file_name)
    ws = wb.active

    # Get column letters
    dataset_col = 1  # First column ("Dataset")
    text_col = 2     # Second column ("Text")

    start_row = 2  # first row of data (row 1 = header)
    current_dataset = ws.cell(row=start_row, column=dataset_col).value
    merge_start = start_row

    for row in range(start_row, ws.max_row + 1):
        dataset_value = ws.cell(row=row, column=dataset_col).value

        if dataset_value != current_dataset:
            # Merge previous dataset cells
            ws.merge_cells(
                start_row=merge_start, start_column=dataset_col,
                end_row=row - 1, end_column=dataset_col
            )
            # Center align the merged cell
            ws.cell(row=merge_start, column=dataset_col).alignment = ws.cell(row=merge_start, column=dataset_col).alignment.copy(horizontal="center", vertical="center")

            # Reset for new dataset
            current_dataset = dataset_value
            merge_start = row

    # Merge the last dataset group
    ws.merge_cells(
        start_row=merge_start, start_column=dataset_col,
        end_row=ws.max_row, end_column=dataset_col
    )
    ws.cell(row=merge_start, column=dataset_col).alignment = ws.cell(row=merge_start, column=dataset_col).alignment.copy(horizontal="center", vertical="center")

    # Save
    wb.save(file_name)

    print("✅ Excel created with merged dataset name cells:", file_name)


def main():
    from glob import glob
    splits=glob("/data/projects/13003558/zoux/datasets/datasets_hf_stage_AudioLLM_v3/datasets_multimodal/train/ASR/*_hok_*")
    print(len(splits), flush=True)
    splits.sort()
    samples_dict={}
    for split in splits:
        ds_name               = os.path.basename(split)
        samples               = generate(split)
        samples_dict[ds_name] = samples
    save_path = "/data/projects/13003558/zoux/datasets/samples/datasets_hf_stage_AudioLLM_v3/datasets_multimodal/train/ASR_hok.json"
    with open(save_path, "w", encoding="utf8") as f_out:
        f_out.write(json.dumps(samples_dict, ensure_ascii=False, indent=2))
    json2excel(save_path)


if __name__ == "__main__":
    main()
